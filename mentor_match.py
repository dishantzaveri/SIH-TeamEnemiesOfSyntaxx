from openpyxl import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from mentee import *
from utils import *
from mentor import *
from re import *
import sys
from string import *

def main():
    wb = Workbook()
    mentee = load_workbook('mentees2020.xlsx')
    mentors = load_workbook('mentors2020.xlsx')
    #mentorFemale = load_workbook('mentorInfoFemale.xlsx')
    matches = wb.active
    matches.title = "Mentor to Mentee"
    menteeSheet = mentee.active
    mentorSheet = mentors.active
    #mentorFemaleSheet = mentorFemale.active
    row = 2
    allMentees = []
    allMentors = []
    while menteeSheet.cell(row=row, column=3).value!=None:
        menteeRow = processMentee(menteeSheet, row)
        mentee = Mentee(row, menteeRow[0], menteeRow[1], menteeRow[2], menteeRow[3], menteeRow[4], menteeRow[5], menteeRow[6], menteeRow[7], menteeRow[8], menteeRow[9], "", 00.00)
        allMentees.append(mentee)
        row = row + 1
    row = 2
    while mentorSheet.cell(row=row, column=2).value!=None:
        mentorRow = processMentor(mentorSheet, row)
        mentor = Mentor(row, mentorRow[0], mentorRow[1], mentorRow[2], mentorRow[3], mentorRow[4], mentorRow[5], mentorRow[6], mentorRow[7], mentorRow[8], 0, [])
        row = row + 1
        allMentors.append(mentor)
    row = 2
    for mentee in allMentees:
        #mentee.printAll()
        mentorMatch(mentee, allMentors)
    
    #Statistics
    #Average Match
    sumofMatches = float(0)
    for mentee in allMentees:
        sumofMatches += mentee.matchRate
    print "Average Mentor Match: " + str(round(sumofMatches/len(allMentees), 2)*100) +"%"

    allMentors.sort(key=lambda x: x.lastName, reverse=False)
    row = 1
    matches.cell(row=row, column=1).value = "Mentors"
    matches.cell(row=row, column=1).font = Font(bold=True, size=12)
    matches.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    matches.cell(row=row, column=2).value = "Mentees"
    matches.cell(row=row, column=2).font = Font(bold=True, size=12)
    matches.cell(row=row, column=2).alignment = Alignment(horizontal="left")
    matches.cell(row=row, column=3).value = "Gender"
    matches.cell(row=row, column=3).font = Font(bold=True, size=12)
    matches.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    matches.cell(row=row, column=4).value = "Race"
    matches.cell(row=row, column=4).font = Font(bold=True, size=12)
    matches.cell(row=row, column=4).alignment = Alignment(horizontal="left")
    matches.cell(row=row, column=5).value = "Grade"
    matches.cell(row=row, column=5).font = Font(bold=True, size=12)
    matches.cell(row=row, column=5).alignment = Alignment(horizontal="left")
    matches.cell(row=row, column=6).value = "Middle School"
    matches.cell(row=row, column=6).font = Font(bold=True, size=12)
    matches.cell(row=row, column=6).alignment = Alignment(horizontal="left")
    matches.cell(row=row, column=7).value = "Match"
    matches.cell(row=row, column=7).font = Font(bold=True, size=12)
    matches.cell(row=row, column=7).alignment = Alignment(horizontal="right")
    matches.cell(row=row, column=8).value = "Interests"
    matches.cell(row=row, column=8).font = Font(bold=True, size=12)
    matches.cell(row=row, column=8).alignment = Alignment(horizontal="left")
    matches.cell(row=row, column=9).value = "Home Phone"
    matches.cell(row=row, column=9).font = Font(bold=True, size=12)
    matches.cell(row=row, column=9).alignment = Alignment(horizontal="left")
    matches.cell(row=row, column=10).value = "Cell Phone"
    matches.cell(row=row, column=10).font = Font(bold=True, size=12)
    matches.cell(row=row, column=10).alignment = Alignment(horizontal="left")
    matches.cell(row=row, column=11).value = "Email"
    matches.cell(row=row, column=11).font = Font(bold=True, size=12)
    matches.cell(row=row, column=11).alignment = Alignment(horizontal="left")
    row = 2
    for mentor in allMentors:
        matches.cell(row=row, column=1).value = mentor.firstName + " " + mentor.lastName
        matches.cell(row=row, column=1).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=2).value = mentor.menteeCount
        matches.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        matches.cell(row=row, column=2).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=3).value = mentor.gender
        matches.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        matches.cell(row=row, column=3).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=4).value = mentor.race
        matches.cell(row=row, column=4).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=5).value = mentor.grade
        matches.cell(row=row, column=5).alignment = Alignment(horizontal="right")
        matches.cell(row=row, column=5).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=6).value = ""
        matches.cell(row=row, column=6).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=7).value = ""
        matches.cell(row=row, column=7).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=8).value = join(mentor.interests, ", ")
        matches.cell(row=row, column=8).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=9).value = ""
        matches.cell(row=row, column=9).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=10).value = mentor.phone
        matches.cell(row=row, column=10).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=11).value = mentor.email
        matches.cell(row=row, column=11).fill = PatternFill("solid", fgColor="DDDDDD")
        row += 1
        for mentee in allMentees:
            if mentee.mentor == mentor.firstName + " " + mentor.lastName:
                if mentee.emailType == "Student":
                    emailType = "(S)"
                else:
                    emailType = "(P)"
                matches.cell(row=row, column=2).value = mentee.firstName + " " + mentee.lastName
                matches.cell(row=row, column=3).value = mentee.gender
                matches.cell(row=row, column=3).alignment = Alignment(horizontal="center")
                matches.cell(row=row, column=4).value = mentee.race
                matches.cell(row=row, column=5).value = "8"
                matches.cell(row=row, column=5).alignment = Alignment(horizontal="right")
                matches.cell(row=row, column=6).value = mentee.middleSchool
                matches.cell(row=row, column=7).value = str(round(mentee.matchRate*100,2))+"%"
                matches.cell(row=row, column=7).alignment = Alignment(horizontal="right")
                matches.cell(row=row, column=8).value = join(mentee.interests, ", ")
                matches.cell(row=row, column=9).value = mentee.homePhone
                matches.cell(row=row, column=10).value = mentee.studentCell
                matches.cell(row=row, column=11).value = mentee.email + " " + emailType
                row += 1
            else:
                row = row
    matches.cell(row=row, column=1).value = "No Interests Given"
    matches.cell(row=row, column=1).fill = PatternFill("solid", fgColor="DDDDDD")
    matches.cell(row=row, column=2).value = ""
    matches.cell(row=row, column=2).fill = PatternFill("solid", fgColor="DDDDDD")
    matches.cell(row=row, column=3).value = ""
    matches.cell(row=row, column=3).fill = PatternFill("solid", fgColor="DDDDDD")
    matches.cell(row=row, column=4).value = ""
    matches.cell(row=row, column=4).fill = PatternFill("solid", fgColor="DDDDDD")
    matches.cell(row=row, column=5).value = ""
    matches.cell(row=row, column=5).fill = PatternFill("solid", fgColor="DDDDDD")
    matches.cell(row=row, column=6).value = ""
    matches.cell(row=row, column=6).fill = PatternFill("solid", fgColor="DDDDDD")
    matches.cell(row=row, column=7).value = ""
    matches.cell(row=row, column=7).fill = PatternFill("solid", fgColor="DDDDDD")
    row+=1
    for mentee in allMentees:
        if mentee.mentor == "No Match":
            if mentee.mentor == mentor.firstName + " " + mentor.lastName:
                if mentee.emailType == "Student":
                    emailType = "(S)"
                else:
                    emailType = "(P)"
            matches.cell(row=row, column=2).value = mentee.firstName + " " + mentee.lastName
            matches.cell(row=row, column=3).value = mentee.gender
            matches.cell(row=row, column=3).alignment = Alignment(horizontal="center")
            matches.cell(row=row, column=4).value = mentee.race
            matches.cell(row=row, column=5).value = "8"
            matches.cell(row=row, column=5).alignment = Alignment(horizontal="right")
            matches.cell(row=row, column=6).value = mentee.middleSchool
            matches.cell(row=row, column=7).value = ""
            matches.cell(row=row, column=8).value = ""
            matches.cell(row=row, column=9).value = mentee.homePhone
            matches.cell(row=row, column=10).value = mentee.studentCell
            matches.cell(row=row, column=11).value = mentee.email + " " + emailType
            row += 1
        else:
            row = row
    wb.save("MentorMenteeMatch.xlsx")

if __name__ == '__main__':
    main()